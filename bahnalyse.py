import requests
import json
import xml.etree.ElementTree as ET
import http.client
import xml.dom.minidom
import openpyxl
import datetime
from openpyxl import load_workbook
from datetime import date
from datetime import datetime
import csv
import os.path

final_planned_data = []
csv_headings_plan_data = ["ID", "Linie","Abfahrt Plan", "Richtung", "Gleis", "Abfragezeitpunkt"]
csv_headings_change_data = ["ID", "Abfahrt laut Change XML", "Verspätung", "Abfragezeitpunkt"]

# EVA number of the station of interest 
stationnr = "8003735"  # Lohhof

#TODO for users: request credentials for DB API MArketplace for personal USE
db_client_id = "xyz"
db_api_key = "123"



def timestamp_diff(start_timestamp, end_timestamp):
    start_year = int(start_timestamp[:2])
    start_month = int(start_timestamp[2:4])
    start_day = int(start_timestamp[4:6])
    start_hour = int(start_timestamp[6:8])
    start_minute = int(start_timestamp[8:])

    end_year = int(end_timestamp[:2])
    end_month = int(end_timestamp[2:4])
    end_day = int(end_timestamp[4:6])
    end_hour = int(end_timestamp[6:8])
    end_minute = int(end_timestamp[8:])

    start_datetime = datetime(start_year, start_month, start_day, start_hour, start_minute)
    end_datetime = datetime(end_year, end_month, end_day, end_hour, end_minute)

    time_diff = end_datetime - start_datetime
    diff_minutes = int(time_diff.total_seconds() / 60)

    return diff_minutes



def export_to_csv(filename, data, headings):
    if os.path.isfile(filename):
        # file exists, append data
        with open(filename, 'a', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for row in data:
                writer.writerow(row)
    else:
        # file doesn't exist, write new data with header row
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headings)
            for row in data:
                writer.writerow(row)
                

                
def export_to_excel(filename, data, headings):
    if os.path.isfile(filename):
        # file exists, append data
        workbook = load_workbook(filename)
        worksheet = workbook.active
        for row in data:
            worksheet.append(row)
    else:
        # file doesn't exist, write new data with header row
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(headings)
        for row in data:
            worksheet.append(row)
    
    workbook.save(filename)


    
# von dem XML einer einzigen Stunde die geplanten Abfahrtszeiten extrahieren
# returns a list of tuples. 
def extract_plan_data_from_xml(xml_plan):
    planned_data = []
    root = ET.fromstring(xml_plan)
    
    for s in root:
        s_id = s.get("id")
        category = s.find("tl").get("c")
        line = s.find("dp").get("l")
        planned_time = s.find("dp").get("pt")
        planned_path = s.find("dp").get("ppth")
        destination = planned_path[:planned_path.find("|")] #next stop
        gleis = s.find("dp").get("pp")
    
        record = (s_id, category+line, planned_time, destination, "Gleis " + gleis, timestamp_plan_request)
        planned_data.append(record)
    
    # optional: nach Richtung München filtern:
    planned_data_uheim = []
    for stop in planned_data:
        if stop[3] == "Unterschleißheim":
            planned_data_uheim.append(stop)
            
    # optional: Duplikate entfernen, weil jeder Zugteil einzeln gewertet wird (zumindest wenn Zugteilung erfolgt)
    planned_data_no_duplicates = []
    
    for stop in planned_data_uheim:
        if len(planned_data_no_duplicates) == 0:
            planned_data_no_duplicates.append(stop)
        else:
            duplicate = False
            for stop_new in planned_data_no_duplicates:
                if stop[1] == stop_new[1] and stop[2] == stop_new[2] and stop[3] == stop_new[3] and stop[4] == stop_new[4]:
                    duplicate = True
            if duplicate == False:
                planned_data_no_duplicates.append(stop)
            
            
    # optional: nach Abfahrtszeit sortieren
    sorted_data = sorted(planned_data_no_duplicates, key=lambda x: x[2])
    return sorted_data



# --------------------------- extract, transform and store plan data (1x taeglich reicht) ---------------------------

# Zeitstempel der Abfrage erstellen
now = datetime.now()
timestamp_plan_request = now.strftime("%Y-%m-%d %H:%M:%S")

conn = http.client.HTTPSConnection("apis.deutschebahn.com")

headers = {
    'DB-Client-Id': db_client_id,
    'DB-Api-Key': db_api_key,
    'accept': "application/xml"
    }

# TODO: adjust this if run daily on server
date = "230408"

#hours = ["00", "01","02", "03","04", "05","06", "07", "08", "09", "10", "11", "12","13", "14", "15","16", "17", "18", "19", "20", "21", "22", "23"]

# for testing, use less requests
hours = ["06", "07","08"]

plan_xmls = []

for hour in hours:
    conn.request("GET", "/db-api-marketplace/apis/timetables/v1/plan/" + stationnr + "/" + date + "/" + hour, headers=headers)
    res = conn.getresponse()
    data = res.read()
    plan_xmls.append(data)


day_data = []

for xml in plan_xmls:
    # skip the given hour if there is no information 
    if len(xml) < 10:
        continue
    else:
        day_data = day_data + extract_plan_data_from_xml(xml)



export_to_csv('planned_data.csv', day_data, csv_headings_plan_data)

# optional/additional: store as xlsx
export_to_excel('planned_data.xlsx', day_data, csv_headings_plan_data)



# --------------------------- extract, transform and store change data (optimalerweise alle 20 min abfragen) ---------------------------

# Die IDs aller Abfahrten die uns interessieren
ids = []
for x in day_data:
    ids.append(x[0])


# Zeitstempel der Abfrage erstellen
now = datetime.now()
timestamp_change_request = now.strftime("%Y-%m-%d %H:%M:%S")

conn = http.client.HTTPSConnection("apis.deutschebahn.com")

headers = {
    'DB-Client-Id': db_client_id,
    'DB-Api-Key': db_api_key,
    'accept': "application/xml"
    }

# full changes 
conn.request("GET", "https://apis.deutschebahn.com/db-api-marketplace/apis/timetables/v1/fchg/8003735", headers=headers)

# recent changes 
#conn.request("GET", "https://apis.deutschebahn.com/db-api-marketplace/apis/timetables/v1/rchg/8003735", headers=headers)

res = conn.getresponse()
data = res.read()


change_data = []

xml_change = data
root = ET.fromstring(xml_change)

for s in root:
    s_id = s.get("id")
    
    #wir holen usn nur die ID die uns interessieren (Richtung Unterschleißheim, keine doppelten)
    if not s_id in ids:
        continue
    
    changed_time = s.find("dp").get("ct")
    
    # delay rausfinden
    for planned_stop in day_data:
        if s_id == planned_stop[0]:
            delay = timestamp_diff(planned_stop[2], changed_time)
    
    record = (s_id, changed_time, delay, timestamp_change_request)
    change_data.append(record)
    
# nach change time sortieren
change_data.sort(key=lambda x: x[1])



export_to_csv('change_data.csv', change_data, csv_headings_change_data)

# optional/additional: store as xlsx
export_to_excel('change_data.xlsx', change_data, csv_headings_change_data)